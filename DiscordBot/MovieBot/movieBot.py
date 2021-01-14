# bot.py
import random
from openpyxl import load_workbook
import time
import os
import discord
from dotenv import load_dotenv
from discord.ext import commands
import asyncio
import urllib.request
from bs4 import BeautifulSoup


class Node(object):
  def __init__(self, name, next):
    self.movieName = name
    self.nextNode = next

  def setName(self, name):
    self.movieName = name

  def setNext(self, next):
    self.nextNode = next

load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')

client = discord.Client()

@client.event
async def on_message(message):
    if message.content.lower() == 'choose movie':
        await message.channel.send('Choosing Movie...')

    if message.author == client.user:
        start = message.content.split(' ')[0]
        if start == 'Choosing':
            moviesWorkBook = load_workbook(filename = 'Movies.xlsx')
            moviesSheet = moviesWorkBook.active

            firstNode = Node("",None)
            currentNode = Node("",None)
            prevNode = Node("",None)


            count = 0
            for currRow in range(1,int(moviesSheet.max_row) + 2): # Two here to ensure empty cell found
              if moviesSheet.cell(row=currRow, column = 1).value == None:
                prevNode.setNext(firstNode)
                break

              if not moviesSheet.cell(row=currRow, column = 3).value == None:
                continue

              if count == 0:
                firstNode.setName(moviesSheet.cell(row=currRow, column = 1).value)
                prevNode = firstNode
                count += 1
                continue

              currentNode.setName(moviesSheet.cell(row=currRow, column = 1).value)
              prevNode.setNext(currentNode)
              prevNode = currentNode
              currentNode = Node("",None)
              count += 1

            randPick1 = random.randint(1, count) # Where to start in the movie list
            randPick2 =  random.randint(10, 15) # How many movies to move through
            total = randPick2
            origTotal = total
            toShow = []

            currentNode = firstNode
            for i in range(0,randPick1):
                currentNode = currentNode.nextNode


            while not total == 0:
              total -= 1
              totalRatio = total/origTotal

              temp = currentNode
              for i in range(0,9):
                toShow.append(temp.movieName)
                temp = temp.nextNode
              printCount = 0
              if total == 1:
                  toPrint = 'Results\n'
              else:
                  toPrint = 'Spinning Wheel...\n'
              toPrint += '\n------------------------------\n'
              for name in toShow:
                if printCount == 4:
                    toPrint+='\n'
                    toPrint+='>>>'+name+'\n'
                    toPrint+='\n'
                else:
                    toPrint+=name+'\n'
                printCount += 1
              toPrint += '------------------------------\n'
              toShow=[]
              currentNode = currentNode.nextNode
              await message.edit(content=toPrint)
              await asyncio.sleep(.8)






client.run(TOKEN)
